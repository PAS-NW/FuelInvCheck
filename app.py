import io
import re
import zipfile
from pathlib import Path
from datetime import datetime
from urllib.parse import quote
from html import escape
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st

try:
    from pypdf import PdfReader
except Exception:
    try:
        from PyPDF2 import PdfReader
    except Exception:
        PdfReader = None

PAS_YELLOW = "#FFD400"
PAS_BLACK = "#0A0A0A"
PAS_DARK = "#171717"
PAS_GREY = "#F4F4F4"

st.set_page_config(page_title="PAS Fuel Invoice Matching", page_icon="pas_logo.png", layout="wide")

st.markdown(
    f"""
    <style>
    .stApp {{ background: #f5f5f5; color: #0A0A0A; }}
    section[data-testid="stSidebar"] {{
        background: {PAS_BLACK};
        color: white;
        padding-top: 1.45rem;
    }}
    section[data-testid="stSidebar"] * {{ color: white; }}
    section[data-testid="stSidebar"] img {{
        margin-top: 0.15rem;
        border-radius: 14px;
    }}
    .block-container {{
        padding-top: 1.4rem;
        padding-bottom: 2rem;
        max-width: 1500px;
    }}

    .pas-hero {{
        background: linear-gradient(135deg, {PAS_BLACK} 0%, #202020 70%, #7a6900 135%);
        border-radius: 18px;
        padding: 24px 28px;
        margin-bottom: 18px;
        box-shadow: 0 8px 25px rgba(0,0,0,0.12);
    }}
    .pas-title {{
        color: white;
        font-size: 32px;
        font-weight: 900;
        margin: 0;
        letter-spacing: -0.03em;
    }}
    .pas-subtitle {{
        color: {PAS_YELLOW};
        font-size: 14px;
        margin-top: 4px;
        font-weight: 800;
    }}

    .kpi-card {{
        background: white;
        border-radius: 18px;
        padding: 18px 20px;
        border: 1px solid #e8e8e8;
        box-shadow: 0 3px 12px rgba(0,0,0,0.05);
        min-height: 112px;
    }}
    .kpi-label {{
        color: #111;
        font-size: 14px;
        font-weight: 800;
        margin-bottom: 8px;
    }}
    .kpi-value {{
        color: {PAS_YELLOW};
        font-size: 36px;
        font-weight: 950;
        line-height: 1.05;
        text-shadow: 0 1px 0 #111;
    }}
    .kpi-sub {{
        color: #222;
        font-size: 13px;
        margin-top: 6px;
    }}

    .stButton > button, .stDownloadButton > button {{
        background: {PAS_YELLOW} !important;
        color: {PAS_BLACK} !important;
        border: 1px solid {PAS_BLACK} !important;
        border-radius: 12px !important;
        font-weight: 900 !important;
    }}

    /* Keep app helper text readable */
    .stCaption, div[data-testid="stCaptionContainer"], .stMarkdown p, .stInfo {{
        color: #0A0A0A !important;
    }}

    .pas-results-title {{
        color: #0A0A0A;
        font-size: 26px;
        font-weight: 950;
        margin: 22px 0 8px 0;
    }}
    .pas-unmatched-pill {{
        display: inline-flex;
        align-items: center;
        gap: 8px;
        background: {PAS_YELLOW};
        color: {PAS_BLACK};
        border: 1px solid #111;
        border-radius: 14px 14px 0 0;
        padding: 11px 18px;
        font-weight: 950;
        box-shadow: 0 3px 10px rgba(0,0,0,0.08);
        margin-top: 4px;
    }}

    .pas-table-wrap {{
        background: white;
        border: 1px solid #d9d9d9;
        border-radius: 0 16px 16px 16px;
        overflow: auto;
        box-shadow: 0 4px 18px rgba(0,0,0,0.07);
        margin-bottom: 18px;
    }}
    table.pas-table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 13px;
        color: #0A0A0A;
        background: white;
    }}
    table.pas-table thead th {{
        background: {PAS_YELLOW};
        color: {PAS_BLACK};
        font-weight: 950;
        text-align: left;
        padding: 11px 12px;
        border: 1px solid #c7a900;
        white-space: nowrap;
    }}
    table.pas-table tbody td {{
        background: white;
        color: #0A0A0A;
        padding: 9px 12px;
        border: 1px solid #e3e3e3;
        vertical-align: top;
    }}
    table.pas-table tbody tr:nth-child(even) td {{
        background: #fbfbfb;
    }}
    table.pas-table a {{
        color: #006fd6 !important;
        font-weight: 800;
        text-decoration: none;
    }}
    table.pas-table a:hover {{
        text-decoration: underline;
    }}
    .pas-note {{
        color: #0A0A0A;
        font-size: 13px;
        margin: 8px 0 16px 0;
    }}
    .pas-support {{
        color: #0A0A0A;
        font-size: 14px;
        margin: 16px 0;
    }}
    .pas-support a {{
        color: #006fd6 !important;
        font-weight: 800;
    }}


    /* --- hard hide Streamlit's uploaded-file chip/list while keeping uploader button usable --- */
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"],
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFileName"],
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFileSize"],
    div[data-testid="stFileUploader"] ul,
    div[data-testid="stFileUploader"] div[role="list"],
    div[data-testid="stFileUploader"] div[role="listitem"] {{
        display: none !important;
        visibility: hidden !important;
        height: 0 !important;
        min-height: 0 !important;
        margin: 0 !important;
        padding: 0 !important;
        overflow: hidden !important;
    }}
    div[data-testid="stFileUploader"] div:has(button[title*="Remove"]),
    div[data-testid="stFileUploader"] div:has(button[aria-label*="Remove"]),
    div[data-testid="stFileUploader"] div:has(svg[data-testid="DeleteIcon"]) {{
        display: none !important;
    }}
    div[data-testid="stFileUploader"] section > div:not(:has(button)) {{
        display: none !important;
    }}
    div[data-testid="stFileUploader"] button {{
        background: #ffffff !important;
        color: #0A0A0A !important;
        border: 1px solid #d7dce3 !important;
        border-radius: 10px !important;
        font-weight: 900 !important;
        box-shadow: 0 2px 8px rgba(0,0,0,.06) !important;
    }}
    div[data-testid="stFileUploader"] button * {{ color:#0A0A0A !important; fill:#0A0A0A !important; stroke:#0A0A0A !important; }}

    </style>
    """,
    unsafe_allow_html=True,
)


st.markdown(
    """
    <style>
    /* Keep sidebar readable on black */
    section[data-testid="stSidebar"] .stMarkdown p,
    section[data-testid="stSidebar"] .stMarkdown li,
    section[data-testid="stSidebar"] .stMarkdown h1,
    section[data-testid="stSidebar"] .stMarkdown h2,
    section[data-testid="stSidebar"] .stMarkdown h3,
    section[data-testid="stSidebar"] .stMarkdown strong,
    section[data-testid="stSidebar"] .stMarkdown span {
        color: #ffffff !important;
    }

    /* Make upload icons visible on dark bars */
    div[data-testid="stFileUploader"] svg,
    div[data-testid="stFileUploader"] button svg,
    div[data-testid="stFileUploader"] [data-testid="stIconMaterial"] {
        color: #FFD400 !important;
        fill: #FFD400 !important;
        stroke: #FFD400 !important;
    }
    div[data-testid="stFileUploader"] section {
        background: #24242d !important;
        border: 1px solid #30303a !important;
        border-radius: 12px !important;
    }
    div[data-testid="stFileUploader"] button {
        color: white !important;
        border-color: #454552 !important;
        background: #111217 !important;
    }

    /* Results table: white body, yellow sticky header, 10-row scroll area */
    .pas-table-wrap {
        max-height: 510px !important;
        overflow-y: auto !important;
        overflow-x: auto !important;
    }
    .pas-table-wrap thead th {
        position: sticky;
        top: 0;
        z-index: 2;
    }
    .pas-note, .pas-support, .pas-support * {
        color: #0A0A0A !important;
    }

    /* Bottom chase animation: small, low, runs once */
    .pas-bottom-chase-wrap {
        position: fixed;
        left: calc(18rem + 22px);
        right: 42px;
        bottom: 12px;
        height: 58px;
        pointer-events: none;
        z-index: 1;
        overflow: hidden;
    }
    .pas-bottom-ground {
        position: absolute;
        left: 0;
        right: 0;
        bottom: 6px;
        border-bottom: 1px solid rgba(0,0,0,0.11);
    }
    .pas-chase-pack {
        position: absolute;
        bottom: 8px;
        left: -150px;
        width: 150px;
        height: 48px;
        animation: pas-chase-run 13s linear 1 forwards;
    }
    @keyframes pas-chase-run {
        0% { transform: translateX(-120px); opacity: 0; }
        8% { opacity: 1; }
        88% { opacity: 1; }
        100% { transform: translateX(calc(100vw - 90px)); opacity: 0; }
    }
    .pas-truck-mini {
        position: absolute;
        left: 0;
        bottom: 5px;
        width: 54px;
        height: 30px;
        filter: drop-shadow(0 1px 1px rgba(0,0,0,.22));
    }
    .pas-truck-bed {
        position: absolute;
        left: 0;
        top: 5px;
        width: 34px;
        height: 19px;
        background: #FFD400;
        border: 3px solid #0A0A0A;
        border-radius: 4px 2px 3px 5px;
        transform: skewX(-10deg);
    }
    .pas-truck-logo {
        position: absolute;
        left: 7px;
        top: 9px;
        font-size: 9px;
        font-weight: 950;
        color: #0A0A0A;
        line-height: 1;
        z-index: 3;
    }
    .pas-truck-cab {
        position: absolute;
        left: 30px;
        top: 7px;
        width: 19px;
        height: 18px;
        background: #FFD400;
        border: 3px solid #0A0A0A;
        border-radius: 3px 5px 3px 2px;
        z-index: 2;
    }
    .pas-truck-window {
        position: absolute;
        left: 34px;
        top: 10px;
        width: 7px;
        height: 7px;
        background: #a8d8e8;
        border: 2px solid #0A0A0A;
        border-radius: 2px;
        z-index: 4;
    }
    .pas-truck-nose {
        position: absolute;
        left: 47px;
        top: 17px;
        width: 8px;
        height: 8px;
        background: #FFD400;
        border: 3px solid #0A0A0A;
        border-left: none;
        border-radius: 0 3px 3px 0;
    }
    .pas-wheel {
        position: absolute;
        bottom: 0;
        width: 9px;
        height: 9px;
        background: #0A0A0A;
        border: 2px solid #222;
        border-radius: 50%;
        animation: pas-wheel-spin .32s linear infinite;
        z-index: 5;
    }
    .pas-wheel::after {
        content: "";
        position: absolute;
        inset: 2px;
        background: #FFD400;
        border-radius: 50%;
    }
    .pas-wheel.back { left: 13px; }
    .pas-wheel.front { left: 41px; }
    @keyframes pas-wheel-spin { to { transform: rotate(360deg); } }

    .pas-speed-lines { position: absolute; left: -30px; top: 17px; width: 24px; height: 18px; }
    .pas-speed-lines span { display:block; height:2px; background:#b9b9b9; margin:4px 0; border-radius:2px; animation: pas-flicker .55s linear infinite; }
    .pas-speed-lines span:nth-child(2) { width: 16px; margin-left: 8px; }
    .pas-speed-lines span:nth-child(3) { width: 11px; margin-left: 13px; }
    @keyframes pas-flicker { 50% { opacity:.25; transform: translateX(-5px); } }

    .pas-dust { position:absolute; left:-5px; bottom:0; width:34px; height:14px; opacity:.75; }
    .pas-dust span { position:absolute; bottom:0; background:#dac6a9; border-radius:50%; animation: pas-dust 1s linear infinite; }
    .pas-dust span:nth-child(1) { width:12px; height:6px; left:0; }
    .pas-dust span:nth-child(2) { width:16px; height:7px; left:10px; animation-delay:.2s; }
    .pas-dust span:nth-child(3) { width:11px; height:5px; left:23px; animation-delay:.4s; }
    @keyframes pas-dust { 50% { transform: translateX(-8px) scale(1.15); opacity:.4; } }

    .pas-stickman {
        position: absolute;
        left: 92px;
        bottom: 5px;
        width: 28px;
        height: 34px;
        animation: pas-runner-bob .35s ease-in-out infinite alternate;
    }
    @keyframes pas-runner-bob { from { transform: translateY(1px); } to { transform: translateY(-2px); } }
    .pas-stick-head {
        position:absolute;
        top:0;
        left:11px;
        width:8px;
        height:8px;
        border:2px solid #111;
        border-radius:50%;
        background:white;
    }
    .pas-stick-body { position:absolute; left:15px; top:9px; width:2px; height:13px; background:#111; transform: rotate(12deg); transform-origin:top; }
    .pas-stick-arm-a, .pas-stick-arm-b, .pas-stick-leg-a, .pas-stick-leg-b { position:absolute; width:2px; height:12px; background:#111; transform-origin:top; border-radius:2px; }
    .pas-stick-arm-a { left:15px; top:11px; transform: rotate(58deg); animation: pas-arm-a .35s linear infinite alternate; }
    .pas-stick-arm-b { left:15px; top:11px; transform: rotate(-50deg); animation: pas-arm-b .35s linear infinite alternate; }
    .pas-stick-leg-a { left:16px; top:21px; height:14px; transform: rotate(48deg); animation: pas-leg-a .35s linear infinite alternate; }
    .pas-stick-leg-b { left:16px; top:21px; height:14px; transform: rotate(-42deg); animation: pas-leg-b .35s linear infinite alternate; }
    @keyframes pas-arm-a { to { transform: rotate(-45deg); } }
    @keyframes pas-arm-b { to { transform: rotate(55deg); } }
    @keyframes pas-leg-a { to { transform: rotate(-45deg); } }
    @keyframes pas-leg-b { to { transform: rotate(48deg); } }
    </style>
    """,
    unsafe_allow_html=True,
)


st.markdown(
    f"""
    <style>
    /* ===== PAS V2 target layout overrides: safe Streamlit-native controls ===== */
    .stApp {{ background: #f7f8fa !important; color: #0A0A0A !important; font-family: Inter, "Segoe UI", Arial, sans-serif; }}
    .block-container {{ max-width: 1580px !important; padding-top: 1.05rem !important; padding-left: 2rem !important; padding-right: 2rem !important; padding-bottom: 2rem !important; }}

    section[data-testid="stSidebar"] {{ background: linear-gradient(180deg, #050606 0%, #0b1015 100%) !important; border-right: 1px solid #161b22; }}
    section[data-testid="stSidebar"] > div:first-child {{ padding-top: 1.05rem !important; }}
    section[data-testid="stSidebar"] img {{ border-radius: 14px !important; box-shadow: 0 10px 24px rgba(0,0,0,.26); }}
    .pas-sidebar-title {{ color:#fff; font-size:18px; font-weight:950; line-height:1.15; text-align:center; margin: 20px 0 8px; }}
    .pas-yellow-line {{ width:72px; height:4px; background:{PAS_YELLOW}; border-radius:99px; margin: 0 auto 22px; }}
    .pas-sidebar-copy {{ color:#fff !important; font-size:14px; line-height:1.52; font-weight:650; margin-bottom:24px; }}
    .pas-sidebar-rule {{ border-top:1px solid rgba(255,255,255,.22); margin:22px 0; }}
    .pas-sidebar-heading {{ color:{PAS_YELLOW}; font-size:19px; font-weight:950; margin: 0 0 16px; }}
    .pas-nav-row {{ display:grid; grid-template-columns: 26px 1fr; gap:10px; align-items:start; margin: 15px 0; color:#fff; font-weight:750; line-height:1.25; font-size:14px; }}
    .pas-nav-icon svg {{ width:21px; height:21px; stroke:{PAS_YELLOW}; stroke-width:2.4; fill:none; stroke-linecap:round; stroke-linejoin:round; }}
    .pas-sidebar-footer {{ color:#fff; font-size:12px; font-weight:800; margin-top:28px; }}

    .pas-hero {{ display:flex; align-items:center; gap:16px; background: linear-gradient(100deg, #08090b 0%, #151718 70%, #c9aa00 130%) !important; border-radius: 16px !important; padding: 12px 22px !important; margin: 0 0 18px 0 !important; box-shadow: 0 9px 25px rgba(0,0,0,.13) !important; min-height:60px; }}
    .pas-hero-logo {{ width:37px; height:37px; border-radius:7px; background:{PAS_YELLOW}; color:#000; display:inline-flex; align-items:center; justify-content:center; font-weight:950; font-size:14px; letter-spacing:-1px; }}
    .pas-hero-text {{ color:#fff; font-size:18px; font-weight:950; letter-spacing:-.02em; }}
    .pas-hero-dot {{ color:#fff; opacity:.8; margin: 0 7px; }}
    .pas-hero-version {{ color:{PAS_YELLOW}; font-weight:950; }}

    .pas-upload-card {{ background:#fff; border:1px solid #e5e7eb; border-radius:18px; box-shadow:0 5px 18px rgba(15,23,42,.08); padding:16px 18px 14px; margin-bottom:14px; }}
    .pas-upload-title {{ color:#0A0A0A; font-size:16px; font-weight:950; margin-bottom:10px; }}
    div[data-testid="stFileUploader"] {{ margin:0 !important; }}
    div[data-testid="stFileUploader"] label {{ display:none !important; }}
    div[data-testid="stFileUploader"] section {{ background:#f4f6f8 !important; border:1px solid #dfe4ea !important; border-radius:11px !important; min-height:52px !important; padding:8px 10px !important; }}
    div[data-testid="stFileUploader"] section * {{ color:#0A0A0A !important; }}
    div[data-testid="stFileUploader"] button {{ background:#fff !important; color:#0A0A0A !important; border:1px solid #d7dce3 !important; border-radius:10px !important; font-weight:900 !important; box-shadow:0 2px 8px rgba(0,0,0,.06) !important; }}
    div[data-testid="stFileUploader"] svg {{ color:#0A0A0A !important; fill:currentColor !important; stroke:currentColor !important; }}
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] {{ background:#fff !important; border:1px solid #dfe4ea !important; border-radius:10px !important; color:#0A0A0A !important; }}
    div[data-testid="stFileUploader"] small {{ color:#4b5563 !important; }}

    div.stButton > button[kind="secondary"], .stButton > button {{ min-height:52px !important; font-size:16px !important; box-shadow:0 6px 18px rgba(255,212,0,.25) !important; }}
    .stDownloadButton > button {{ min-height:62px !important; font-size:20px !important; box-shadow:0 6px 18px rgba(255,212,0,.25) !important; }}

    .kpi-card {{ background:#fff !important; border-radius:18px !important; border:1px solid #e4e7eb !important; box-shadow:0 5px 20px rgba(15,23,42,.08) !important; min-height:118px !important; padding:18px 22px !important; display:flex; align-items:center; gap:18px; }}
    .kpi-icon {{ width:64px; height:64px; border-radius:50%; background:#fff5bd; display:flex; align-items:center; justify-content:center; flex:none; }}
    .kpi-icon svg {{ width:35px; height:35px; stroke:#0A0A0A; stroke-width:2.5; fill:none; stroke-linecap:round; stroke-linejoin:round; }}
    .kpi-label {{ color:#111 !important; font-size:15px !important; font-weight:950 !important; margin:0 0 3px !important; }}
    .kpi-value {{ color:#e9b900 !important; font-size:42px !important; line-height:.98 !important; font-weight:950 !important; text-shadow:none !important; }}
    .kpi-sub {{ color:#374151 !important; font-size:14px !important; margin-top:6px !important; }}
    .kpi-unmatched .kpi-value {{ color:#e12626 !important; }}
    .kpi-matched .kpi-value {{ color:#16a34a !important; }}

    .pas-results-title {{ color:#0A0A0A !important; font-size:28px !important; font-weight:950 !important; margin: 22px 0 8px !important; }}
    .pas-unmatched-pill {{ background:{PAS_YELLOW} !important; color:#0A0A0A !important; border:0 !important; border-radius:14px 14px 0 0 !important; padding:13px 20px !important; font-size:18px; box-shadow:0 4px 14px rgba(0,0,0,.09); }}
    .pas-table-wrap {{ background:#fff !important; border:1px solid #e0e4e9 !important; border-radius:0 16px 16px 16px !important; max-height:430px !important; overflow:auto !important; box-shadow:0 7px 25px rgba(15,23,42,.10) !important; }}
    table.pas-table {{ font-size:14px !important; color:#0A0A0A !important; }}
    table.pas-table thead th {{ background:{PAS_YELLOW} !important; color:#0A0A0A !important; border:1px solid #e2ba00 !important; padding:12px 14px !important; font-weight:950 !important; position:sticky; top:0; z-index:5; }}
    table.pas-table tbody td {{ background:#fff !important; color:#0A0A0A !important; border:1px solid #e1e5eb !important; padding:10px 14px !important; }}
    table.pas-table tbody tr:nth-child(even) td {{ background:#fbfcfd !important; }}
    .pas-pdf-icon {{ display:inline-flex; align-items:center; justify-content:center; width:17px; height:20px; background:#e11d2e; color:#fff; border-radius:3px; font-size:9px; font-weight:950; margin-right:8px; vertical-align:middle; }}
    table.pas-table a {{ color:#006bd6 !important; font-weight:850 !important; }}
    table.pas-table .query-cell {{ min-width:120px; white-space:nowrap; }}
    .pas-note, .pas-support, .pas-support * {{ color:#0A0A0A !important; }}
    .pas-support {{ margin-top:22px !important; font-size:15px !important; }}
    .pas-support a {{ color:#006bd6 !important; font-weight:900 !important; margin-left:12px; }}

    /* --- uploader chip cleanup: hide Streamlit's ugly uploaded-file pill and use our own card --- */
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] {{ display: none !important; }}
    div[data-testid="stFileUploaderDropzone"] {{ background: transparent !important; border: 0 !important; padding: 0 !important; min-height: 0 !important; }}
    div[data-testid="stFileUploaderDropzoneInstructions"] {{ display: none !important; }}
    div[data-testid="stFileUploader"] section {{ background: transparent !important; border: 0 !important; min-height: 0 !important; padding: 0 !important; }}
    div[data-testid="stFileUploader"] button {{
        background: #ffffff !important;
        color: #0A0A0A !important;
        border: 1px solid #d7dce3 !important;
        border-radius: 10px !important;
        font-weight: 900 !important;
        min-height: 44px !important;
        box-shadow: 0 2px 8px rgba(0,0,0,.06) !important;
    }}
    .pas-file-card {{
        display:flex; align-items:center; gap:14px;
        background:#f4f6f8; border:1px solid #dfe4ea; border-radius:12px;
        padding:11px 14px; min-height:54px; margin: 4px 0 12px;
    }}
    .pas-file-icon {{ width:32px; height:32px; border-radius:8px; display:flex; align-items:center; justify-content:center; color:#fff; font-weight:950; font-size:11px; box-shadow:0 2px 8px rgba(0,0,0,.12); flex:none; }}
    .pas-file-icon.excel {{ background:#118a3b; }}
    .pas-file-icon.pdf {{ background:#df1f2d; }}
    .pas-file-main {{ flex:1; min-width:0; }}
    .pas-file-name {{ color:#0A0A0A; font-weight:950; font-size:15px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
    .pas-file-size {{ color:#4b5563; font-weight:650; font-size:13px; margin-top:2px; }}
    .pas-file-check {{ width:24px; height:24px; border-radius:50%; background:#108a37; color:white; display:flex; align-items:center; justify-content:center; font-size:15px; font-weight:950; flex:none; }}
    </style>
    """,
    unsafe_allow_html=True,
)

with st.sidebar:
    st.image("pas_logo.png", use_column_width=True)
    st.markdown(
        """
        <div class="pas-sidebar-title">PAS Fuel<br>Invoice Matching</div>
        <div class="pas-yellow-line"></div>
        <div class="pas-sidebar-copy">Upload the Vehicle spreadsheet and fuel invoice PDF, then export annotated PDF and Excel summary.</div>
        <div class="pas-sidebar-rule"></div>
        <div class="pas-sidebar-heading">Instructions</div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M16 16l-4-4-4 4"/><path d="M12 12v9"/><path d="M20 16.6A5 5 0 0 0 18 7h-1.3A8 8 0 1 0 4 15.3"/></svg></span><span>Upload Vehicle Spreadsheet</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><path d="M14 2v6h6"/><path d="M9 13h6"/><path d="M9 17h6"/></svg></span><span>Upload Fuel Invoice PDF</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M5 3l14 9-14 9V3z"/></svg></span><span>Run Reconciliation</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><path d="M7 10l5 5 5-5"/><path d="M12 15V3"/></svg></span><span>Download Reconciliation<br>PDF & Excel</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><circle cx="11" cy="11" r="8"/><path d="M21 21l-4.3-4.3"/></svg></span><span>Smoke Crack</span></div>
        <div class="pas-sidebar-rule"></div>
        <div class="pas-sidebar-footer">PAS NW Ltd • v1.0.1 Weekend Date Fix</div>
        """,
        unsafe_allow_html=True,
    )

st.markdown(
    """
    <div class="pas-hero">
      <div class="pas-hero-logo">PAS</div>
      <div class="pas-hero-text">PAS NW Ltd<span class="pas-hero-dot">•</span><span class="pas-hero-version">v1.0.1 Weekend Date Fix</span></div>
    </div>
    """,
    unsafe_allow_html=True,
)



def render_bottom_chase():
    """Small non-intrusive PAS dump truck chase animation pinned to the bottom of the white content area."""
    st.markdown(
        """
        <div class="pas-bottom-chase-wrap" aria-hidden="true">
            <div class="pas-bottom-ground"></div>
            <div class="pas-chase-pack">
                <div class="pas-speed-lines"><span></span><span></span><span></span></div>
                <div class="pas-dust"><span></span><span></span><span></span></div>
                <div class="pas-truck-mini">
                    <div class="pas-truck-bed"></div>
                    <div class="pas-truck-logo">PAS</div>
                    <div class="pas-truck-cab"></div>
                    <div class="pas-truck-window"></div>
                    <div class="pas-truck-nose"></div>
                    <div class="pas-wheel back"></div>
                    <div class="pas-wheel front"></div>
                </div>
                <div class="pas-stickman">
                    <div class="pas-stick-head"></div>
                    <div class="pas-stick-body"></div>
                    <div class="pas-stick-arm-a"></div>
                    <div class="pas-stick-arm-b"></div>
                    <div class="pas-stick-leg-a"></div>
                    <div class="pas-stick-leg-b"></div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )




st.markdown(
    """
    <style>
    div[data-testid="stAlert"], div[data-testid="stAlert"] * {
        color: #0A0A0A !important;
    }
    div[data-testid="stAlert"] {
        border: 1px solid #e2ba00 !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ===== Fuel invoice reconciliation logic =====
try:
    from pypdf import PdfWriter
except Exception:
    try:
        from PyPDF2 import PdfWriter
    except Exception:
        PdfWriter = None

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.colors import black
except Exception:
    canvas = None

try:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except Exception:
    PatternFill = Font = Alignment = Border = Side = get_column_letter = None

FUEL_RESULT_COLUMNS = [
    "Date",
    "Card No",
    "Vehicle Registration",
    "Driver Name",
    "Job Number / Site",
    "Status",
]

FUEL_EXCEL_COLUMNS = [
    "Date",
    "Card No",
    "Vehicle Registration",
    "Driver Name",
    "Job Number / Site",
    "Fuel Quantity",
    "Fuel Value",
    "Status",
    "Page",
    "Transaction Detail",
]


def render_selected_file_card(uploaded_file, file_kind="excel"):
    size = getattr(uploaded_file, "size", 0) or 0
    if size >= 1024 * 1024:
        size_text = f"{size / (1024 * 1024):.1f} MB"
    else:
        size_text = f"{size / 1024:.0f} KB"
    icon = "XLS" if file_kind == "excel" else "PDF"
    icon_class = "excel" if file_kind == "excel" else "pdf"
    st.markdown(
        f'''
        <div class="pas-file-card">
            <div class="pas-file-icon {icon_class}">{icon}</div>
            <div class="pas-file-main">
                <div class="pas-file-name">{escape(getattr(uploaded_file, "name", "Uploaded file"))}</div>
                <div class="pas-file-size">{size_text}</div>
            </div>
            <div class="pas-file-check">✓</div>
        </div>
        ''',
        unsafe_allow_html=True,
    )


def clean_cell(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    return text


def normalise_card(value) -> str:
    # Keep only numbers for fuel-card matching.
    # Motia invoices show the short card number, while Vehicles.xlsx stores the
    # long card number with a leading * and often a trailing check digit.
    return re.sub(r"[^0-9]+", "", clean_cell(value)).lstrip("0")


def fuel_card_matches(invoice_card: str, spreadsheet_card: str) -> bool:
    inv = normalise_card(invoice_card)
    sheet = normalise_card(spreadsheet_card)
    if not inv or not sheet:
        return False
    if inv == sheet:
        return True
    # Vehicles.xlsx stores cards like *7055059234413344690 while the invoice
    # shows 334469. Match the invoice short card inside the long card number.
    if len(inv) >= 5 and inv in sheet:
        return True
    # Safety fallback for cards where only the last 6 digits are comparable.
    if len(inv) >= 5 and len(sheet) >= len(inv) and sheet[-len(inv):] == inv:
        return True
    return False


def find_col(columns: List[str], keywords: List[str]) -> Optional[str]:
    norm_cols = {c: re.sub(r"[^a-z0-9]+", "", str(c).lower()) for c in columns}
    for key in keywords:
        nkey = re.sub(r"[^a-z0-9]+", "", key.lower())
        for col, ncol in norm_cols.items():
            if nkey == ncol:
                return col
    for key in keywords:
        nkey = re.sub(r"[^a-z0-9]+", "", key.lower())
        for col, ncol in norm_cols.items():
            if nkey in ncol:
                return col
    return None


def money_to_float(value) -> Optional[float]:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).replace(",", "")
    matches = re.findall(r"-?£?\s*([0-9]+(?:\.[0-9]{2})?)", text)
    if not matches:
        return None
    try:
        return round(float(matches[-1]), 2)
    except Exception:
        return None


def load_vehicles_workbook(uploaded_file) -> Dict[str, List[Dict[str, str]]]:
    xls = pd.ExcelFile(uploaded_file)
    sheet_lookup = {s.lower().strip(): s for s in xls.sheet_names}
    vehicles_sheet = sheet_lookup.get("vehicles") or xls.sheet_names[0]
    sold_sheet = sheet_lookup.get("off hire - sold") or sheet_lookup.get("off hire sold")

    def read_sheet(sheet_name: str, source_priority: int) -> List[Dict[str, str]]:
        if not sheet_name:
            return []
        df = pd.read_excel(xls, sheet_name=sheet_name).dropna(how="all")
        cols = list(df.columns)
        card_col = find_col(cols, ["Fuel Card No", "Fuel Card Number", "Fuel Card", "Card No", "Card Number"])
        driver_col = find_col(cols, ["Driver Name", "Driver", "Employee", "User"])
        site_col = find_col(cols, ["Job Number / Site", "Job No", "Job", "Site", "Project", "Location"])
        reg_col = find_col(cols, ["Vehicle Registration", "Registration", "Reg", "VRN", "Vehicle Reg"])
        if not card_col:
            return []
        out = []
        for _, row in df.iterrows():
            raw_card = clean_cell(row.get(card_col, ""))
            card = normalise_card(raw_card)
            if not card or card in {"-", "0"}:
                continue
            out.append({
                "Fuel Card Raw": raw_card,
                "Fuel Card Normalised": card,
                "Driver Name": clean_cell(row.get(driver_col, "")) if driver_col else "",
                "Job Number / Site": clean_cell(row.get(site_col, "")) if site_col else "",
                "Vehicle Registration": clean_cell(row.get(reg_col, "")) if reg_col else "",
                "Source Priority": source_priority,  # 1 = Vehicles, 2 = Off Hire - Sold
                "Source Sheet": sheet_name,
            })
        return out

    vehicle_rows = read_sheet(vehicles_sheet, 1)
    sold_rows = read_sheet(sold_sheet, 2) if sold_sheet else []
    return {"rows": vehicle_rows + sold_rows}


def find_vehicle_for_card(invoice_card: str, card_lookup: Dict[str, List[Dict[str, str]]]) -> Dict[str, str]:
    rows = card_lookup.get("rows", []) if isinstance(card_lookup, dict) else []
    matches = [r for r in rows if fuel_card_matches(invoice_card, r.get("Fuel Card Normalised", ""))]
    if not matches:
        return {}
    # Priority rule: Vehicles tab always wins over Off Hire - Sold.
    matches.sort(key=lambda r: r.get("Source Priority", 99))
    return matches[0]


def read_pdf_pages(uploaded_pdf) -> Tuple[List[str], bytes]:
    pdf_bytes = uploaded_pdf.read()
    uploaded_pdf.seek(0)
    if PdfReader is None:
        raise RuntimeError("PDF reader is unavailable. Add pypdf to requirements.txt.")
    reader = PdfReader(io.BytesIO(pdf_bytes))
    pages = []
    for page in reader.pages:
        try:
            pages.append(page.extract_text() or "")
        except Exception:
            pages.append("")
    return pages, pdf_bytes


def extract_fuel_transactions_from_text(pages: List[str]) -> List[Dict[str, str]]:
    """Extract Motia / Fuel Card Services transaction rows from invoice text.

    The Motia invoice uses dates like 20-May-26 and transaction rows such as:
    20-May-26 06:29 334469 SHELL SALE 071603 MJ71HKB MK66HPV 105438 Diesel 57.78 ...

    Site names can contain spaces and occasionally wrap onto the next line, so the
    extraction anchors on the fixed fields: date, time, card number, TXN number,
    vehicle registration, odometer, product, quantity and net value.
    """
    transactions = []
    date_token = r"\d{1,2}[-/][A-Za-z]{3}[-/]\d{2,4}\*?|\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\*?"
    time_token = r"\d{1,2}:\d{2}"
    reg_token = r"[A-Z]{1,3}\d{1,3}[A-Z]{0,3}|[A-Z]{2}\d{2}[A-Z]{3}"
    product_token = r"Diesel|Unleaded(?:\s+Medium(?:\s+Octane)?)?|Petrol|AdBlue"

    for page_no, text in enumerate(pages, start=1):
        lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines()]
        i = 0
        while i < len(lines):
            raw = lines[i]
            if not raw:
                i += 1
                continue

            low = raw.lower()
            if "annual card fee" in low:
                card_match = re.search(r"\b(\d{5,20})\b", raw)
                transactions.append({
                    "Date": "",
                    "Card No": card_match.group(1) if card_match else "",
                    "Vehicle Registration": "",
                    "Fuel Quantity": "",
                    "Fuel Value": "",
                    "Raw Line": raw,
                    "Page": page_no,
                })
                i += 1
                continue

            if not re.match(rf"^{date_token}\s+{time_token}\s+\d{{5,20}}\b", raw, re.I):
                i += 1
                continue

            # Some site names wrap, e.g. 'M65 DARWEN MOTORWAY' / 'SERVICE' / transaction tail.
            # Keep joining continuation lines until the numeric tail is present or a new record/total starts.
            parse_line = raw
            lookahead = i + 1
            while lookahead < len(lines) and lookahead <= i + 4:
                nxt = lines[lookahead]
                if not nxt:
                    lookahead += 1
                    continue
                if re.match(rf"^{date_token}\s+{time_token}\s+\d{{5,20}}\b", nxt, re.I):
                    break
                if nxt.upper().startswith("CARD SUB TOTAL") or nxt.upper().startswith("CUSTOMER TOTAL") or nxt.lower().startswith("continued"):
                    break
                parse_line += " " + nxt
                if re.search(rf"({product_token})\s+\d+(?:\.\d+)?\s+\d+(?:\.\d+)?\s+\d+(?:\.\d+)?\s+\d+(?:\.\d+)?\s+\d+(?:\.\d+)?\s+\d+(?:\.\d+)?\b", parse_line, re.I):
                    break
                lookahead += 1

            # Prefer Motia/Fuel Card Services layout.
            pattern = re.compile(
                rf"^(?P<date>{date_token})\s+"
                rf"(?P<time>{time_token})\s+"
                rf"(?P<card>\d{{5,20}})\s+"
                rf"(?P<site>.+?)\s+"
                rf"(?P<txn>\d{{5,12}})\s+"
                rf"(?P<vehreg>{reg_token})\s+"
                rf"(?P<cardtext>.+?)\s+"
                rf"(?P<odo>\d{{1,9}})\s+"
                rf"(?P<product>{product_token})\s+"
                rf"(?P<qty>\d+(?:\.\d+)?)\s+"
                rf"(?P<unit>\d+(?:\.\d+)?)\s+"
                rf"(?P<net>\d+(?:\.\d+)?)\s+"
                rf"(?P<vatrate>\d+(?:\.\d+)?)\s+"
                rf"(?P<vat>\d+(?:\.\d+)?)\s+"
                rf"(?P<gross>\d+(?:\.\d+)?)\b",
                re.I,
            )
            m = pattern.search(parse_line)
            if m:
                transactions.append({
                    "Date": m.group("date"),
                    "Card No": m.group("card"),
                    "Vehicle Registration": m.group("vehreg").upper(),
                    "Fuel Quantity": m.group("qty"),
                    "Fuel Value": m.group("net"),
                    "Raw Line": parse_line,
                    "Page": page_no,
                })
                i += 1
                continue

            # Fallback for other fuel-card formats: parse from fixed leading fields,
            # but choose the first long number after time as the card number rather than
            # the last number on the line.
            m_head = re.match(rf"^(?P<date>{date_token})\s+{time_token}\s+(?P<card>\d{{5,20}})\b", raw, re.I)
            if m_head:
                reg = ""
                reg_match = re.search(rf"\b({reg_token})\b", raw, re.I)
                if reg_match:
                    reg = reg_match.group(1).upper()
                nums = re.findall(r"\b\d+(?:\.\d+)?\b", raw)
                qty = nums[-6] if len(nums) >= 6 else (nums[-5] if len(nums) >= 5 else "")
                net = nums[-4] if len(nums) >= 6 else (nums[-3] if len(nums) >= 3 else "")
                transactions.append({
                    "Date": m_head.group("date"),
                    "Card No": m_head.group("card"),
                    "Vehicle Registration": reg,
                    "Fuel Quantity": qty,
                    "Fuel Value": net,
                    "Raw Line": raw,
                    "Page": page_no,
                })
            i += 1

    seen = set()
    out = []
    for tx in transactions:
        key = (tx.get("Date"), tx.get("Card No"), tx.get("Fuel Value"), tx.get("Raw Line"))
        if key in seen:
            continue
        seen.add(key)
        out.append(tx)
    return out

def reconcile_fuel_transactions(transactions: List[Dict[str, str]], card_lookup: Dict[str, Dict[str, str]]) -> pd.DataFrame:
    rows = []
    for tx in transactions:
        raw_line = tx.get("Raw Line", "")
        card_no = tx.get("Card No", "")
        vehicle = find_vehicle_for_card(card_no, card_lookup) if card_no else {}
        if "annual card fee" in raw_line.lower():
            driver = "Card Renewal"
            site = "Card Renewal"
            status = "Card Renewal"
        elif vehicle:
            driver = vehicle.get("Driver Name", "")
            site = vehicle.get("Job Number / Site", "")
            status = "Matched"
        else:
            driver = ""
            site = ""
            status = "Unmatched"
        rows.append({
            "Date": tx.get("Date", ""),
            "Card No": card_no,
            "Vehicle Registration": tx.get("Vehicle Registration", "") or vehicle.get("Vehicle Registration", ""),
            "Driver Name": driver,
            "Job Number / Site": site,
            "Fuel Quantity": tx.get("Fuel Quantity", ""),
            "Fuel Value": tx.get("Fuel Value", ""),
            "Status": status,
            "Page": tx.get("Page", 1),
            "Raw Line": raw_line,
        })
    return pd.DataFrame(rows)


def style_excel(writer, dfs: Dict[str, pd.DataFrame]):
    if not all([PatternFill, Font, Alignment, Border, Side, get_column_letter]):
        return
    yellow = PatternFill("solid", fgColor="FFD400")
    header_font = Font(name="Calibri", size=10, bold=True, color="000000")
    body_font = Font(name="Calibri", size=10, color="000000")
    alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"), top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
    for sheet_name, df in dfs.items():
        ws = writer.book[sheet_name]
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 25
        for row in ws.iter_rows():
            for cell in row:
                cell.font = header_font if cell.row == 1 else body_font
                cell.alignment = alignment
                cell.border = border
                if cell.row == 1:
                    cell.fill = yellow
        ws.auto_filter.ref = ws.dimensions
        for idx, col in enumerate(df.columns, start=1):
            values = [str(col)] + [str(v) for v in df[col].fillna("").astype(str).tolist()]
            width = min(max(len(v) for v in values) + 2, 45)
            ws.column_dimensions[get_column_letter(idx)].width = width


def make_job_cost_summary(reconciliation_df: pd.DataFrame) -> pd.DataFrame:
    """Build the Excel summary tab: total fuel cost by job number/site."""
    if reconciliation_df is None or reconciliation_df.empty:
        return pd.DataFrame(columns=["Job Number / Site", "Total Fuel Cost", "Transaction Count", "Driver Names"])

    df = reconciliation_df.copy()
    df["Job Number / Site"] = df.get("Job Number / Site", "").apply(clean_cell)
    df["Job Number / Site"] = df["Job Number / Site"].replace("", "Unmatched / No Job Number")
    df["Fuel Value Numeric"] = df.get("Fuel Value", "").apply(lambda v: money_to_float(v) or 0.0)

    summary = (
        df.groupby("Job Number / Site", dropna=False)
        .agg(
            **{
                "Total Fuel Cost": ("Fuel Value Numeric", "sum"),
                "Transaction Count": ("Fuel Value Numeric", "size"),
                "Driver Names": ("Driver Name", lambda s: ", ".join(sorted({clean_cell(v) for v in s if clean_cell(v)}))),
            }
        )
        .reset_index()
        .sort_values("Job Number / Site", key=lambda s: s.astype(str).str.lower())
    )
    summary["Total Fuel Cost"] = summary["Total Fuel Cost"].round(2)
    return summary


def make_fuel_excel(run_summary_df: pd.DataFrame, reconciliation_df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    export_df = reconciliation_df.copy()
    if "Raw Line" in export_df.columns and "Transaction Detail" not in export_df.columns:
        export_df["Transaction Detail"] = export_df["Raw Line"]
    for col in FUEL_EXCEL_COLUMNS:
        if col not in export_df.columns:
            export_df[col] = ""
    export_df = export_df[FUEL_EXCEL_COLUMNS]

    job_summary_df = make_job_cost_summary(reconciliation_df)
    dfs = {
        "Job Summary": job_summary_df,
        "Transactions": export_df,
        "Run Summary": run_summary_df,
    }
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet, df in dfs.items():
            df.to_excel(writer, index=False, sheet_name=sheet)
        style_excel(writer, dfs)
    return output.getvalue()


def make_annotated_pdf(original_pdf_bytes: bytes, reconciliation_df: pd.DataFrame) -> bytes:
    """Create a clean side-note PDF.

    The original invoice page is kept intact. A narrow right-hand margin is added
    and each matched transaction gets a Driver | Job note aligned to its row.
    Unmatched rows are intentionally left blank.
    """
    try:
        import fitz  # PyMuPDF
    except Exception:
        return original_pdf_bytes

    try:
        doc = fitz.open(stream=original_pdf_bytes, filetype="pdf")
        date_re = re.compile(r"^\d{1,2}[-/][A-Za-z]{3}[-/]\d{2,4}\*?$|^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\*?$", re.I)
        time_re = re.compile(r"^\d{1,2}:\d{2}$")
        card_re = re.compile(r"^\d{5,20}$")
        margin_width = 155

        for page_number, page in enumerate(doc, start=1):
            page_rows = reconciliation_df[reconciliation_df.get("Page", 1) == page_number]
            if page_rows.empty:
                continue

            old_rect = page.rect
            old_width = old_rect.width
            old_height = old_rect.height
            new_width = old_width + margin_width
            page.set_mediabox(fitz.Rect(0, 0, new_width, old_height))

            # White side-note area with a subtle divider so the invoice itself remains untouched.
            margin_rect = fitz.Rect(old_width, 0, new_width, old_height)
            page.draw_rect(margin_rect, color=None, fill=(1, 1, 1), overlay=True)
            page.draw_line((old_width + 2, 0), (old_width + 2, old_height), color=(0.84, 0.84, 0.84), width=0.5, overlay=True)

            words = page.get_text("words")
            # word tuple: x0, y0, x1, y1, word, block_no, line_no, word_no
            words = sorted(words, key=lambda w: (round(w[1], 1), w[0]))
            anchors = []
            for i in range(len(words) - 2):
                w0 = clean_cell(words[i][4])
                w1 = clean_cell(words[i + 1][4])
                w2 = clean_cell(words[i + 2][4])
                if date_re.match(w0) and time_re.match(w1) and card_re.match(w2):
                    anchors.append({
                        "card": w2,
                        "y": (float(words[i][1]) + float(words[i][3])) / 2,
                        "used": False,
                    })

            def note_for_row(row) -> str:
                status = clean_cell(row.get("Status", ""))
                if status == "Unmatched":
                    return ""
                if status == "Card Renewal":
                    return "Card Renewal"
                driver = clean_cell(row.get("Driver Name", ""))
                site = clean_cell(row.get("Job Number / Site", ""))
                return f"{driver} | {site}".strip(" |")

            for _, row in page_rows.iterrows():
                note = note_for_row(row)
                if not note:
                    continue

                card = clean_cell(row.get("Card No", ""))
                chosen = None
                for anchor in anchors:
                    if not anchor["used"] and anchor["card"] == card:
                        chosen = anchor
                        break
                if chosen is None:
                    for anchor in anchors:
                        if not anchor["used"]:
                            chosen = anchor
                            break
                if chosen is None:
                    continue

                chosen["used"] = True
                y = chosen["y"] + 3.0
                text_box = fitz.Rect(old_width + 8, y - 8, new_width - 5, y + 8)
                # Font kept small enough to avoid wrapping on most names, but readable.
                page.insert_textbox(
                    text_box,
                    note[:42],
                    fontsize=7.6,
                    fontname="helv",
                    color=(0, 0, 0),
                    align=fitz.TEXT_ALIGN_LEFT,
                    overlay=True,
                )

        out = io.BytesIO()
        doc.save(out, garbage=4, deflate=True)
        doc.close()
        return out.getvalue()
    except Exception:
        return original_pdf_bytes


def render_results_table(df: pd.DataFrame):
    if df is None or df.empty:
        st.markdown('<div class="pas-unmatched-pill">Results</div>', unsafe_allow_html=True)
        st.markdown('<div class="pas-table-wrap"><table class="pas-table"><tbody><tr><td>No transactions found.</td></tr></tbody></table></div>', unsafe_allow_html=True)
        return
    display_df = df.copy()
    for col in FUEL_RESULT_COLUMNS:
        if col not in display_df.columns:
            display_df[col] = ""
    display_df = display_df[FUEL_RESULT_COLUMNS]
    rows_html = []
    for _, row in display_df.iterrows():
        cells = "".join(f"<td>{escape(clean_cell(row.get(col, '')))}</td>" for col in display_df.columns)
        rows_html.append(f"<tr>{cells}</tr>")
    header_html = "".join(f"<th>{escape(col)}</th>" for col in display_df.columns)
    st.markdown('<div class="pas-unmatched-pill">Fuel Transactions</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="pas-table-wrap"><table class="pas-table"><thead><tr>{header_html}</tr></thead><tbody>{"".join(rows_html)}</tbody></table></div>', unsafe_allow_html=True)


up_col1, up_col2 = st.columns(2)
with up_col1:
    st.markdown('<div class="pas-upload-card"><div class="pas-upload-title">Upload Vehicles.xlsx</div>', unsafe_allow_html=True)
    vehicles_file = st.file_uploader("Upload Vehicles.xlsx", type=["xlsx", "xls"], label_visibility="collapsed", key="vehicles_upload")
    if vehicles_file:
        render_selected_file_card(vehicles_file, "excel")
    st.markdown('</div>', unsafe_allow_html=True)
with up_col2:
    st.markdown('<div class="pas-upload-card"><div class="pas-upload-title">Upload Fuel Invoice PDF</div>', unsafe_allow_html=True)
    fuel_pdf = st.file_uploader("Upload Fuel Invoice PDF", type=["pdf"], label_visibility="collapsed", key="fuel_pdf_upload")
    if fuel_pdf:
        render_selected_file_card(fuel_pdf, "pdf")
    st.markdown('</div>', unsafe_allow_html=True)

run = st.button("▶  Run reconciliation", use_container_width=True)

if "fuel_reconciliation_results" not in st.session_state:
    st.session_state["fuel_reconciliation_results"] = None

if run:
    if not vehicles_file or not fuel_pdf:
        st.warning("Please upload both Vehicles.xlsx and the fuel invoice PDF.")
        st.stop()
    try:
        with st.spinner("Reading vehicle database..."):
            card_lookup = load_vehicles_workbook(vehicles_file)
        with st.spinner("Reading fuel invoice PDF..."):
            pages, original_pdf_bytes = read_pdf_pages(fuel_pdf)
            transactions = extract_fuel_transactions_from_text(pages)
        with st.spinner("Reconciling fuel transactions..."):
            all_df = reconcile_fuel_transactions(transactions, card_lookup)
            if all_df.empty:
                st.warning("No fuel transactions could be extracted from the PDF.")
                st.stop()

        total = len(all_df)
        matched = int((all_df["Status"] == "Matched").sum())
        unmatched = int((all_df["Status"] == "Unmatched").sum())
        card_renewals = int((all_df["Status"] == "Card Renewal").sum())
        match_pct = round((matched / total) * 100, 1) if total else 0.0

        summary_df = pd.DataFrame({
            "Metric": [
                "Total transactions",
                "Matched transactions",
                "Unmatched transactions",
                "Card renewals",
                "Match percentage",
                "Run date/time",
            ],
            "Value": [
                total,
                matched,
                unmatched,
                card_renewals,
                f"{match_pct}%",
                datetime.now().strftime("%d/%m/%Y %H:%M"),
            ],
        })

        excel_bytes = make_fuel_excel(summary_df, all_df)
        annotated_pdf_bytes = make_annotated_pdf(original_pdf_bytes, all_df)

        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        st.session_state["fuel_reconciliation_results"] = {
            "all_df": all_df,
            "summary_df": summary_df,
            "excel_bytes": excel_bytes,
            "annotated_pdf_bytes": annotated_pdf_bytes,
            "total": total,
            "matched": matched,
            "unmatched": unmatched,
            "match_pct": match_pct,
            "excel_filename": f"PAS_Fuel_Reconciliation_{stamp}.xlsx",
            "pdf_filename": f"PAS_Fuel_Invoice_Annotated_{stamp}.pdf",
        }
    except Exception as e:
        st.error(f"Something went wrong: {e}")
        st.exception(e)

results = st.session_state.get("fuel_reconciliation_results")

if results is not None:
    total = results["total"]
    matched = results["matched"]
    unmatched = results["unmatched"]
    match_pct = results["match_pct"]
    all_df = results["all_df"]

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f'<div class="kpi-card"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M8 7V3h8l4 4v14H6V7z"/><path d="M16 3v5h5"/><path d="M9 13h6"/><path d="M9 17h4"/><path d="M4 7h2v14h12"/></svg></div><div><div class="kpi-label">Total transactions</div><div class="kpi-value">{total}</div><div class="kpi-sub">Detected transactions</div></div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="kpi-card kpi-matched"><div class="kpi-icon"><svg viewBox="0 0 24 24"><circle cx="12" cy="12" r="9"/><path d="M8 12.5l2.7 2.7L16.5 9"/></svg></div><div><div class="kpi-label">Matched transactions</div><div class="kpi-value">{matched}</div><div class="kpi-sub">Approved transactions</div></div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="kpi-card kpi-unmatched"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M12 3l10 18H2L12 3z"/><path d="M12 9v5"/><path d="M12 18h.01"/></svg></div><div><div class="kpi-label">Unmatched transactions</div><div class="kpi-value">{unmatched}</div><div class="kpi-sub">Need review</div></div></div>', unsafe_allow_html=True)
    with c4:
        st.markdown(f'<div class="kpi-card"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M3 20h18"/><path d="M6 16v-4"/><path d="M11 16V8"/><path d="M16 16v-6"/><path d="M19 6l-5 5-3-3-5 5"/></svg></div><div><div class="kpi-label">Match %</div><div class="kpi-value">{match_pct}%</div><div class="kpi-sub">Core KPI</div></div></div>', unsafe_allow_html=True)

    st.markdown('<div class="pas-results-title">Results</div>', unsafe_allow_html=True)
    render_results_table(all_df)

    dl_left, dl_right = st.columns(2)
    with dl_left:
        st.download_button(
            "⬇  Download annotated PDF",
            data=results["annotated_pdf_bytes"],
            file_name=results["pdf_filename"],
            mime="application/pdf",
            use_container_width=True,
        )
    with dl_right:
        st.download_button(
            "⬇  Download Excel output",
            data=results["excel_bytes"],
            file_name=results["excel_filename"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
else:
    st.info("Upload Vehicles.xlsx and the fuel invoice PDF, then click Run reconciliation.")
    render_bottom_chase()
